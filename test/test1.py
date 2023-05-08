import openpyxl
import random

# Создаем новый Excel-файл и выбираем активный лист
workbook = openpyxl.Workbook()
sheet = workbook.active

# Заполняем заголовки столбцов
sheet.cell(row=1, column=1, value='Событие')
sheet.cell(row=1, column=2, value='Исход')

# Генерируем 1000 рандомных бросков кубика и записываем результаты в файл
for i in range(2, 1002):
    dice_roll1 = random.randint(1, 6)
    dice_roll2 = random.randint(1, 6)
    sheet.cell(row=i, column=1, value=f'Бросок {i-1}')
    sheet.cell(row=i, column=2, value=int(f'{dice_roll1}{dice_roll2}'))

# Сохраняем файл
workbook.save(filename='dice.xlsx')
